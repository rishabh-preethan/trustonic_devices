from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time

# Path to your webdriver executable
url = 'https://tienda.movistar.com.pe/catalog/category/view/id/29/?amp%3Butm_campaign=PE_MOVIL_COL-DO-EQUIPOSLIBERADOS-B2C_2021-12-20_WEB_LDS-WEB_0_NA&amp%3Butm_content=FILTRO&amp%3Butm_medium=FILTRO&amp%3Butm_source=CATALOGO-WP&product_list_order=name'
url = "https://tienda.movistar.com.pe/celulares/renovacion?product_list_order=name"
url = "https://tienda.movistar.com.pe/celulares/portabilidad?product_list_order=name"
driver = webdriver.Chrome()
driver.get(url)

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['Name', 'Price'])

# Keep track of the last row in the Excel sheet
row = 2

# Find elements for name and price
while True:
    name_elements = driver.find_elements(By.CLASS_NAME, 'product-item-link')
    price_elements = driver.find_elements(By.CLASS_NAME, 'itemDetail-value-3')

    # Extract text for name and price
    for name_element, price_element in zip(name_elements, price_elements):
        name = name_element.text.strip()
        price = price_element.text.strip()
        print(name, price)
        ws.cell(row, 1, name)
        ws.cell(row, 2, price)
        row += 1

    # Click the next page button if it exists
    try:
        next_page_link = driver.find_element(By.XPATH, '//a[contains(@class, "next")]')
        next_page_link.click()
        time.sleep(2)  # Wait for the next page to load
    except:
        break  # If the next page button is not found, exit the loop

# Save the Excel file
wb.save('movistar_peru_potabilidad.xlsx')

# Close the webdriver
driver.quit()
