from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time

# Path to your webdriver executable
url_template = 'https://www.vodacom.co.za/shopping/products?range=83&isShowQualifying=false&currentPage={}'
driver = webdriver.Chrome()

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['Name', 'Price'])

# Keep track of the last row in the Excel sheet
row = 2

# Iterate through pages
for page_num in range(1, 13):
    url = url_template.format(page_num)
    driver.get(url)
    time.sleep(10)
    # Find elements for name and price
    name_elements = driver.find_elements(By.CLASS_NAME, 'ProductCards_mobile-label__vGNhb')
    price_elements = driver.find_elements(By.CLASS_NAME, 'ProductCards_monthly-mobile-text__XQ9G5')

    # Extract text for name and price
    for name_element, price_element in zip(name_elements, price_elements):
        name = name_element.text.strip()
        price = price_element.text.strip().replace('or for ', '')  # Remove additional text
        print(name, price)
        ws.cell(row, 1, name)
        ws.cell(row, 2, price)
        row += 1

# Save the Excel file
wb.save('vodacom_products.xlsx')

# Close the webdriver
driver.quit()
